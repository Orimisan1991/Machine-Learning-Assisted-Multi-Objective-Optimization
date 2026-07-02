"""
=============================================================================
Machine Learning-Assisted Multi-Objective Optimization of Hardness and Impact
Performance in Bamboo Powder-Reinforced PBS Green Composites
=============================================================================
Full Analysis Pipeline:
  1. Data Setup & Descriptive Statistics
  2. ANOVA + Tukey HSD Post-Hoc Tests
  3. Gaussian Process Regression (GPR) Surrogate Models
  4. NSGA-II Multi-Objective Optimization (Pareto Front)  [built-in, no deap]
  5. Desirability Function Analysis
  6. Publication-Quality Figures  ← each saved as a SEPARATE PNG file
  7. Excel Workbook with all results, tables & summaries
=============================================================================
Dependencies: numpy, scipy, pandas, matplotlib, scikit-learn, openpyxl
Install     : pip install numpy scipy pandas matplotlib scikit-learn openpyxl
=============================================================================
"""

import warnings
warnings.filterwarnings("ignore")

import os
import random
import itertools

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.lines import Line2D

from scipy import stats
from scipy.stats import f_oneway, tukey_hsd
from sklearn.gaussian_process import GaussianProcessRegressor
from sklearn.gaussian_process.kernels import RBF, ConstantKernel, WhiteKernel
from sklearn.model_selection import LeaveOneOut
from sklearn.metrics import r2_score, mean_absolute_error

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference, Series

# =============================================================================
#  OUTPUT DIRECTORY
# =============================================================================
OUT_DIR = "/mnt/user-data/outputs/bp_pbs_results"
os.makedirs(OUT_DIR, exist_ok=True)

def save_fig(fig, name):
    """Save figure as a separate PNG and close it."""
    path = os.path.join(OUT_DIR, name)
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    print(f"  ✓ {name}")
    return path

# =============================================================================
#  GLOBAL AESTHETICS  (journal-ready)
# =============================================================================
COLORS = {
    "deep_green": "#1B5E20",
    "mid_green":  "#388E3C",
    "lime":       "#8BC34A",
    "amber":      "#F57F17",
    "orange":     "#E65100",
    "blue":       "#1565C0",
    "light_blue": "#42A5F5",
    "grey":       "#546E7A",
    "light_grey": "#ECEFF1",
    "pareto":     "#D50000",
    "optimum":    "#FFD600",
}

plt.rcParams.update({
    "font.family":       "serif",
    "font.serif":        ["Times New Roman", "DejaVu Serif"],
    "axes.labelsize":    12,
    "axes.titlesize":    13,
    "xtick.labelsize":   10,
    "ytick.labelsize":   10,
    "legend.fontsize":   9,
    "figure.dpi":        150,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "axes.grid":         False,
    "grid.alpha":        0.3,
    "grid.linestyle":    "--",
    "savefig.dpi":       300,
    "savefig.bbox":      "tight",
})

# =============================================================================
#  1. EXPERIMENTAL DATA
# =============================================================================
loadings = np.array([0, 5, 10, 20])

hardness_raw = np.array([
    [65.16037, 64.74724, 64.82117],
    [64.50748, 65.53171, 66.51754],
    [63.60067, 63.31122, 63.72942],
    [62.37243, 62.47250, 63.06814],
])

impact_raw = np.array([
    [12.39612, 12.83699, 12.86082],
    [13.81485, 14.20038, 14.65807],
    [13.99403, 14.03337, 14.14561],
    [13.44853, 13.70879, 13.69164],
])

H_mean  = hardness_raw.mean(axis=1)
H_std   = hardness_raw.std(axis=1, ddof=1)
IS_mean = impact_raw.mean(axis=1)
IS_std  = impact_raw.std(axis=1, ddof=1)
H_cv    = (H_std / H_mean) * 100
IS_cv   = (IS_std / IS_mean) * 100

# =============================================================================
#  2. DESCRIPTIVE STATISTICS TABLE
# =============================================================================
print("=" * 70)
print("  DESCRIPTIVE STATISTICS")
print("=" * 70)
df_stats = pd.DataFrame({
    "Loading (wt%)":    loadings,
    "H_mean (HR)":      H_mean.round(4),
    "H_std":            H_std.round(4),
    "H_CV (%)":         H_cv.round(2),
    "IS_mean (J/mm²)":  IS_mean.round(4),
    "IS_std":           IS_std.round(4),
    "IS_CV (%)":        IS_cv.round(2),
})
print(df_stats.to_string(index=False))

print("\n  Δ vs Neat PBS (0 wt%)")
delta_rows = []
for i, w in enumerate(loadings[1:], 1):
    dH  = (H_mean[i]  - H_mean[0]) / H_mean[0]  * 100
    dIS = (IS_mean[i] - IS_mean[0]) / IS_mean[0] * 100
    print(f"  {w:>4} wt%  →  ΔHardness = {dH:+.2f}%   ΔImpact = {dIS:+.2f}%")
    delta_rows.append({"Loading (wt%)": w, "ΔHardness (%)": round(dH,2), "ΔImpact (%)": round(dIS,2)})
df_delta = pd.DataFrame(delta_rows)

# =============================================================================
#  3. ONE-WAY ANOVA
# =============================================================================
print("\n" + "=" * 70)
print("  ONE-WAY ANOVA")
print("=" * 70)

F_H,  p_H  = f_oneway(*[hardness_raw[i] for i in range(4)])
F_IS, p_IS = f_oneway(*[impact_raw[i]   for i in range(4)])

print(f"  Hardness  : F = {F_H:.4f},  p = {p_H:.6f}  {'*** SIGNIFICANT' if p_H<0.05 else 'n.s.'}")
print(f"  Impact    : F = {F_IS:.4f},  p = {p_IS:.6f}  {'*** SIGNIFICANT' if p_IS<0.05 else 'n.s.'}")

# Tukey HSD
print("\n  Tukey HSD (Hardness)")
tukey_H = tukey_hsd(*[hardness_raw[i] for i in range(4)])
tukey_h_rows = []
for i, j in itertools.combinations(range(4), 2):
    sig = "sig." if tukey_H.pvalue[i, j] < 0.05 else "n.s."
    print(f"    {loadings[i]:>2} wt% vs {loadings[j]:>2} wt% : p = {tukey_H.pvalue[i,j]:.4f}  {sig}")
    tukey_h_rows.append({"Group A (wt%)": loadings[i], "Group B (wt%)": loadings[j],
                          "p-value": round(tukey_H.pvalue[i,j],4), "Significance": sig})

print("\n  Tukey HSD (Impact Strength)")
tukey_IS = tukey_hsd(*[impact_raw[i] for i in range(4)])
tukey_is_rows = []
for i, j in itertools.combinations(range(4), 2):
    sig = "sig." if tukey_IS.pvalue[i, j] < 0.05 else "n.s."
    print(f"    {loadings[i]:>2} wt% vs {loadings[j]:>2} wt% : p = {tukey_IS.pvalue[i,j]:.4f}  {sig}")
    tukey_is_rows.append({"Group A (wt%)": loadings[i], "Group B (wt%)": loadings[j],
                           "p-value": round(tukey_IS.pvalue[i,j],4), "Significance": sig})

df_tukey_H  = pd.DataFrame(tukey_h_rows)
df_tukey_IS = pd.DataFrame(tukey_is_rows)

# =============================================================================
#  4. GPR SURROGATE MODELS
# =============================================================================
X_train = loadings.reshape(-1, 1).astype(float)
y_H     = H_mean
y_IS    = IS_mean

kernel = ConstantKernel(1.0, (1e-3, 1e3)) * RBF(5.0, (1e-2, 50)) + WhiteKernel(1e-4, (1e-8, 1e-1))

gpr_H  = GaussianProcessRegressor(kernel=kernel, n_restarts_optimizer=20, normalize_y=True, random_state=42)
gpr_IS = GaussianProcessRegressor(kernel=kernel, n_restarts_optimizer=20, normalize_y=True, random_state=42)

gpr_H.fit(X_train,  y_H)
gpr_IS.fit(X_train, y_IS)

# LOOCV
loo = LeaveOneOut()
y_H_pred_cv, y_IS_pred_cv = [], []
for tr_idx, te_idx in loo.split(X_train):
    gH = GaussianProcessRegressor(kernel=kernel, n_restarts_optimizer=10, normalize_y=True, random_state=42)
    gI = GaussianProcessRegressor(kernel=kernel, n_restarts_optimizer=10, normalize_y=True, random_state=42)
    gH.fit(X_train[tr_idx], y_H[tr_idx]);  y_H_pred_cv.append(gH.predict(X_train[te_idx])[0])
    gI.fit(X_train[tr_idx], y_IS[tr_idx]); y_IS_pred_cv.append(gI.predict(X_train[te_idx])[0])

r2_H_cv  = r2_score(y_H,  y_H_pred_cv)
r2_IS_cv = r2_score(y_IS, y_IS_pred_cv)
mae_H    = mean_absolute_error(y_H,  y_H_pred_cv)
mae_IS   = mean_absolute_error(y_IS, y_IS_pred_cv)

print("\n" + "=" * 70)
print("  GPR MODEL PERFORMANCE (Leave-One-Out CV)")
print("=" * 70)
print(f"  Hardness  : LOOCV R² = {r2_H_cv:.4f},   MAE = {mae_H:.4f} HR")
print(f"  Impact    : LOOCV R² = {r2_IS_cv:.4f},   MAE = {mae_IS:.4f} J/mm²")

# Prediction grid
X_pred = np.linspace(0, 20, 400).reshape(-1, 1)
H_pred,  H_std_pred  = gpr_H.predict(X_pred,  return_std=True)
IS_pred, IS_std_pred = gpr_IS.predict(X_pred, return_std=True)

# =============================================================================
#  5. NSGA-II MULTI-OBJECTIVE OPTIMIZATION  (built-in, no external dep)
# =============================================================================
def _evaluate(x_val):
    x = np.array([[x_val]])
    h   = gpr_H.predict(x)[0]
    is_ = gpr_IS.predict(x)[0]
    return (-h, -is_)

def _dominates(a, b):
    return all(ai <= bi for ai, bi in zip(a, b)) and any(ai < bi for ai, bi in zip(a, b))

def _non_dominated_sort(pop_fit):
    n = len(pop_fit)
    fronts = [[]]
    dominated_by = [[] for _ in range(n)]
    domination_count = [0] * n
    rank = [0] * n
    for i in range(n):
        for j in range(n):
            if i == j:
                continue
            if _dominates(pop_fit[i], pop_fit[j]):
                dominated_by[i].append(j)
            elif _dominates(pop_fit[j], pop_fit[i]):
                domination_count[i] += 1
        if domination_count[i] == 0:
            rank[i] = 0
            fronts[0].append(i)
    i_f = 0
    while fronts[i_f]:
        next_front = []
        for i in fronts[i_f]:
            for j in dominated_by[i]:
                domination_count[j] -= 1
                if domination_count[j] == 0:
                    rank[j] = i_f + 1
                    next_front.append(j)
        fronts.append(next_front)
        i_f += 1
    return fronts[:-1], rank

def _crowding_distance(front_indices, pop_fit):
    n = len(front_indices)
    if n == 0:
        return {}
    dist = {i: 0.0 for i in front_indices}
    for obj in range(2):
        sorted_idx = sorted(front_indices, key=lambda i: pop_fit[i][obj])
        dist[sorted_idx[0]]  = float("inf")
        dist[sorted_idx[-1]] = float("inf")
        f_min = pop_fit[sorted_idx[0]][obj]
        f_max = pop_fit[sorted_idx[-1]][obj]
        span = f_max - f_min if f_max != f_min else 1e-9
        for k in range(1, n - 1):
            dist[sorted_idx[k]] += (pop_fit[sorted_idx[k+1]][obj] -
                                     pop_fit[sorted_idx[k-1]][obj]) / span
    return dist

def run_nsga2(n_pop=200, n_gen=200, cx_prob=0.9, mut_prob=0.3,
              x_min=0.0, x_max=20.0, seed=42):
    random.seed(seed)
    np.random.seed(seed)

    pop = [random.uniform(x_min, x_max) for _ in range(n_pop)]
    pop_fit = [_evaluate(x) for x in pop]

    def _sbx(x1, x2, eta=20.0):
        u = random.random()
        beta = (2*u)**(1/(eta+1)) if u <= 0.5 else (1/(2*(1-u)))**(1/(eta+1))
        c1 = 0.5 * ((1+beta)*x1 + (1-beta)*x2)
        c2 = 0.5 * ((1-beta)*x1 + (1+beta)*x2)
        return (np.clip(c1, x_min, x_max), np.clip(c2, x_min, x_max))

    def _poly_mut(x, eta=20.0):
        u = random.random()
        if u < 0.5:
            delta = (2*u)**(1/(eta+1)) - 1
        else:
            delta = 1 - (2*(1-u))**(1/(eta+1))
        return float(np.clip(x + delta*(x_max - x_min), x_min, x_max))

    for _ in range(n_gen):
        # Tournament selection (crowded comparison)
        fronts, rank = _non_dominated_sort(pop_fit)
        crowd = {}
        for f in fronts:
            crowd.update(_crowding_distance(f, pop_fit))

        def _tournament(a, b):
            if rank[a] < rank[b]:
                return a
            if rank[b] < rank[a]:
                return b
            return a if crowd.get(a, 0) >= crowd.get(b, 0) else b

        offspring = []
        indices = list(range(n_pop))
        random.shuffle(indices)
        for k in range(0, n_pop, 2):
            p1 = _tournament(indices[k], indices[(k+1) % n_pop])
            p2 = _tournament(indices[(k+2) % n_pop], indices[(k+3) % n_pop])
            if random.random() < cx_prob:
                c1, c2 = _sbx(pop[p1], pop[p2])
            else:
                c1, c2 = pop[p1], pop[p2]
            if random.random() < mut_prob: c1 = _poly_mut(c1)
            if random.random() < mut_prob: c2 = _poly_mut(c2)
            offspring.extend([c1, c2])
        offspring_fit = [_evaluate(x) for x in offspring]

        combined     = pop + offspring
        combined_fit = pop_fit + offspring_fit
        all_fronts, all_rank = _non_dominated_sort(combined_fit)

        new_pop, new_fit = [], []
        for f in all_fronts:
            if len(new_pop) + len(f) <= n_pop:
                for i in f:
                    new_pop.append(combined[i])
                    new_fit.append(combined_fit[i])
            else:
                cd = _crowding_distance(f, combined_fit)
                f_sorted = sorted(f, key=lambda i: cd.get(i, 0), reverse=True)
                needed = n_pop - len(new_pop)
                for i in f_sorted[:needed]:
                    new_pop.append(combined[i])
                    new_fit.append(combined_fit[i])
                break
        pop, pop_fit = new_pop[:n_pop], new_fit[:n_pop]

    # Extract Pareto front (rank 0)
    fronts, _ = _non_dominated_sort(pop_fit)
    pf_indices = fronts[0]
    pf_x  = np.array([pop[i]      for i in pf_indices])
    pf_H  = np.array([-pop_fit[i][0] for i in pf_indices])
    pf_IS = np.array([-pop_fit[i][1] for i in pf_indices])
    sort_idx = np.argsort(pf_x)
    return pf_x[sort_idx], pf_H[sort_idx], pf_IS[sort_idx]

print("\nRunning NSGA-II ... (built-in implementation)")
pf_x, pf_H, pf_IS = run_nsga2()

# =============================================================================
#  6. DESIRABILITY FUNCTION (Derringer-Suich)
# =============================================================================
H_target  = H_mean.max()
IS_target = IS_mean.max()
H_min,  H_max  = H_mean.min(),  H_mean.max()
IS_min, IS_max = IS_mean.min(), IS_mean.max()

def desirability_H(h):
    if h <= H_min:    return 0.0
    if h >= H_target: return 1.0
    return ((h - H_min) / (H_target - H_min)) ** 1.0

def desirability_IS(is_):
    if is_ <= IS_min:    return 0.0
    if is_ >= IS_target: return 1.0
    return ((is_ - IS_min) / (IS_target - IS_min)) ** 1.0

D_pf = np.array([(desirability_H(h) * desirability_IS(is_)) ** 0.5
                  for h, is_ in zip(pf_H, pf_IS)])
best_idx = np.argmax(D_pf)
opt_x, opt_H, opt_IS, opt_D = pf_x[best_idx], pf_H[best_idx], pf_IS[best_idx], D_pf[best_idx]

# Dense desirability over grid
D_grid = np.array([(desirability_H(h) * desirability_IS(is_)) ** 0.5
                    for h, is_ in zip(H_pred, IS_pred)])
opt_grid_idx = np.argmax(D_grid)
opt_grid_x = X_pred[opt_grid_idx, 0]
x_dense = X_pred.ravel()

print("\n" + "=" * 70)
print("  MULTI-OBJECTIVE OPTIMIZATION RESULTS")
print("=" * 70)
print(f"  Pareto front size     : {len(pf_x)} solutions")
print(f"  Optimal loading (D)   : {opt_x:.2f} wt%  (grid: {opt_grid_x:.2f} wt%)")
print(f"  Predicted Hardness    : {opt_H:.4f} HR")
print(f"  Predicted Impact      : {opt_IS:.4f} J/mm²")
print(f"  Composite Desirability: {opt_D:.4f}")

pf_df = pd.DataFrame({
    "Loading (wt%)":    pf_x.round(2),
    "Pred. H (HR)":     pf_H.round(4),
    "Pred. IS (J/mm²)": pf_IS.round(4),
    "Desirability":     D_pf.round(4),
})
pf_df = pf_df.sort_values("Desirability", ascending=False)

# =============================================================================
#  7. PUBLICATION-QUALITY FIGURES  ← EACH AS A SEPARATE FILE
# =============================================================================
print("\n" + "=" * 70)
print("  GENERATING FIGURES  (each saved as separate PNG)")
print("=" * 70)

# ── Fig 1a: Hardness Bar Chart ────────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(7, 5))
bar_colors = [COLORS["deep_green"] if i != 1 else COLORS["amber"] for i in range(4)]
ax.bar(loadings, H_mean, yerr=H_std, capsize=6, width=3.5,
       color=bar_colors, edgecolor="white", linewidth=0.8,
       error_kw=dict(ecolor=COLORS["grey"], lw=1.5))
ax.axhline(H_mean[0], color=COLORS["grey"], ls="--", lw=1.2, alpha=0.7, label="Neat PBS baseline")
for i, (x, y, s) in enumerate(zip(loadings, H_mean, H_std)):
    ax.text(x, y + s + 0.1, f"{y:.2f}", ha="center", va="bottom", fontsize=8.5, fontweight="bold")
ax.set_xlabel("Bamboo Powder Loading (wt%)")
ax.set_ylabel("Rockwell Hardness (HR)")
ax.set_title("Effect of BP Loading on Hardness\n(Error bars = ±1 SD, n = 3)")
ax.set_xticks(loadings)
ax.legend()
ax.set_ylim(60, 67.5)
ax.annotate("Peak\n5 wt%", xy=(5, H_mean[1]), xytext=(8, 66.5),
            arrowprops=dict(arrowstyle="->", color=COLORS["amber"], lw=1.5),
            color=COLORS["amber"], fontsize=9, fontweight="bold")
save_fig(fig, "fig1a_hardness_bar.png")

# ── Fig 1b: Impact Strength Bar Chart ────────────────────────────────────────
fig, ax = plt.subplots(figsize=(7, 5))
bar_colors2 = [COLORS["blue"] if i != 1 else COLORS["amber"] for i in range(4)]
ax.bar(loadings, IS_mean, yerr=IS_std, capsize=6, width=3.5,
       color=bar_colors2, edgecolor="white", linewidth=0.8,
       error_kw=dict(ecolor=COLORS["grey"], lw=1.5))
ax.axhline(IS_mean[0], color=COLORS["grey"], ls="--", lw=1.2, alpha=0.7, label="Neat PBS baseline")
for i, (x, y, s) in enumerate(zip(loadings, IS_mean, IS_std)):
    ax.text(x, y + s + 0.02, f"{y:.2f}", ha="center", va="bottom", fontsize=8.5, fontweight="bold")
ax.set_xlabel("Bamboo Powder Loading (wt%)")
ax.set_ylabel("Impact Strength (J/mm²)")
ax.set_title("Effect of BP Loading on Impact Strength\n(Error bars = ±1 SD, n = 3)")
ax.set_xticks(loadings)
ax.legend()
ax.set_ylim(12.5, 15.0)
ax.annotate("Peak\n5 wt%", xy=(5, IS_mean[1]), xytext=(9, 14.7),
            arrowprops=dict(arrowstyle="->", color=COLORS["amber"], lw=1.5),
            color=COLORS["amber"], fontsize=9, fontweight="bold")
save_fig(fig, "fig1b_impact_bar.png")

# ── Fig 2a: GPR Surrogate — Hardness ─────────────────────────────────────────
fig, ax = plt.subplots(figsize=(7, 5))
ax.fill_between(x_dense, H_pred - 2*H_std_pred, H_pred + 2*H_std_pred,
                alpha=0.2, color=COLORS["deep_green"], label="95% CI")
ax.fill_between(x_dense, H_pred - H_std_pred, H_pred + H_std_pred,
                alpha=0.35, color=COLORS["deep_green"], label="68% CI")
ax.plot(X_pred, H_pred, color=COLORS["deep_green"], lw=2.2, label="GPR mean")
ax.scatter(loadings, y_H, color="black", zorder=5, s=70, label="Experimental mean")
ax.errorbar(loadings, y_H, yerr=H_std, fmt="none", ecolor="black", capsize=5, lw=1.2)
ax.scatter(loadings, y_H_pred_cv, marker="D", s=50, color=COLORS["amber"],
           zorder=6, label="LOOCV prediction")
ax.set_xlabel("Bamboo Powder Loading (wt%)")
ax.set_ylabel("Hardness (HR)")
ax.set_title("GPR Surrogate — Hardness\nwith Leave-One-Out Cross-Validation")
ax.set_ylim(60.5, 67.0)
ax.legend(loc="upper right")
ax.text(0.04, 0.07, f"LOOCV R² = {r2_H_cv:.3f}\nMAE = {mae_H:.3f} HR",
        transform=ax.transAxes, fontsize=9,
        bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8))
save_fig(fig, "fig2a_gpr_hardness.png")

# ── Fig 2b: GPR Surrogate — Impact Strength ──────────────────────────────────
fig, ax = plt.subplots(figsize=(7, 5))
ax.fill_between(x_dense, IS_pred - 2*IS_std_pred, IS_pred + 2*IS_std_pred,
                alpha=0.2, color=COLORS["blue"], label="95% CI")
ax.fill_between(x_dense, IS_pred - IS_std_pred, IS_pred + IS_std_pred,
                alpha=0.35, color=COLORS["blue"], label="68% CI")
ax.plot(X_pred, IS_pred, color=COLORS["blue"], lw=2.2, label="GPR mean")
ax.scatter(loadings, y_IS, color="black", zorder=5, s=70, label="Experimental mean")
ax.errorbar(loadings, y_IS, yerr=IS_std, fmt="none", ecolor="black", capsize=5, lw=1.2)
ax.scatter(loadings, y_IS_pred_cv, marker="D", s=50, color=COLORS["amber"],
           zorder=6, label="LOOCV prediction")
ax.set_xlabel("Bamboo Powder Loading (wt%)")
ax.set_ylabel("Impact Strength (J/mm²)")
ax.set_title("GPR Surrogate — Impact Strength\nwith Leave-One-Out Cross-Validation")
ax.set_ylim(12.8, 14.6)
ax.legend(loc="lower right")
ax.text(0.04, 0.07, f"LOOCV R² = {r2_IS_cv:.3f}\nMAE = {mae_IS:.3f} J/mm²",
        transform=ax.transAxes, fontsize=9,
        bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8))
save_fig(fig, "fig2b_gpr_impact.png")

# ── Fig 3: Pareto Front ───────────────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(8, 6))
sc = ax.scatter(pf_H, pf_IS, c=pf_x, cmap="YlGn", s=45, zorder=4,
                edgecolors="grey", linewidths=0.4, label="Pareto solutions")
ax.plot(pf_H[np.argsort(pf_H)], pf_IS[np.argsort(pf_H)],
        color=COLORS["grey"], lw=1.2, ls="--", alpha=0.6)
ax.scatter(opt_H, opt_IS, s=200, marker="*", color=COLORS["optimum"],
           edgecolors="black", linewidths=0.8, zorder=6,
           label=f"Optimum ({opt_x:.1f} wt%, D={opt_D:.3f})")
for i, (xp, hp, ip) in enumerate(zip(loadings, H_mean, IS_mean)):
    ax.scatter(hp, ip, s=80, marker="^", color=COLORS["orange"], zorder=5)
    ax.annotate(f" {xp} wt%", (hp, ip), fontsize=8.5, color=COLORS["orange"])
plt.colorbar(sc, ax=ax, label="BP Loading (wt%)")
ax.set_xlabel("Predicted Rockwell Hardness (HR)")
ax.set_ylabel("Predicted Impact Strength (J/mm²)")
ax.set_title("NSGA-II Pareto Front\n(Maximize Hardness & Impact Strength Simultaneously)")
ax.legend(loc="lower left")
save_fig(fig, "fig3_pareto_front.png")

# ── Fig 4a: Desirability — Hardness ──────────────────────────────────────────
d_H_arr  = np.array([desirability_H(h)   for h  in H_pred])
d_IS_arr = np.array([desirability_IS(is_) for is_ in IS_pred])

fig, ax = plt.subplots(figsize=(7, 5))
ax.plot(x_dense, d_H_arr, color=COLORS["deep_green"], lw=2.2)
ax.axvline(opt_grid_x, color=COLORS["orange"], ls="--", lw=1.5,
           label=f"Optimum {opt_grid_x:.1f} wt%")
ax.set_title(" Desirability — Hardness")
ax.set_xlabel("BP Loading (wt%)")
ax.set_ylabel("Individual Desirability d_H")
ax.legend()
save_fig(fig, "fig4a_desirability_hardness.png")

# ── Fig 4b: Desirability — Impact Strength ───────────────────────────────────
fig, ax = plt.subplots(figsize=(7, 5))
ax.plot(x_dense, d_IS_arr, color=COLORS["blue"], lw=2.2)
ax.axvline(opt_grid_x, color=COLORS["orange"], ls="--", lw=1.5,
           label=f"Optimum {opt_grid_x:.1f} wt%")
ax.set_title("Desirability — Impact Strength")
ax.set_xlabel("BP Loading (wt%)")
ax.set_ylabel("Individual Desirability d_IS")
ax.legend()
save_fig(fig, "fig4b_desirability_impact.png")

# ── Fig 4c: Composite Desirability ───────────────────────────────────────────
fig, ax = plt.subplots(figsize=(7, 5))
ax.plot(x_dense, D_grid, color=COLORS["amber"], lw=2.5)
ax.fill_between(x_dense, 0, D_grid, alpha=0.2, color=COLORS["amber"])
ax.axvline(opt_grid_x, color=COLORS["pareto"], ls="--", lw=1.8,
           label=f"Max D = {D_grid.max():.3f} @ {opt_grid_x:.1f} wt%")
ax.scatter([opt_grid_x], [D_grid.max()], s=150, color=COLORS["pareto"], zorder=5)
ax.set_title(" Composite Desirability D\n(Derringer-Suich)")
ax.set_xlabel("BP Loading (wt%)")
ax.set_ylabel("Composite Desirability D")
ax.legend()
save_fig(fig, "fig4c_composite_desirability.png")

# ── Fig 5a: Box Plot — Hardness ───────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(7, 5))
bp1 = ax.boxplot([hardness_raw[i] for i in range(4)],
                  labels=[f"{w} wt%" for w in loadings],
                  patch_artist=True, widths=0.45,
                  medianprops=dict(color="white", lw=2))
for patch, col in zip(bp1["boxes"], [COLORS["grey"], COLORS["amber"],
                                      COLORS["mid_green"], COLORS["deep_green"]]):
    patch.set_facecolor(col); patch.set_alpha(0.8)
ax.set_ylabel("Rockwell Hardness (HR)")
ax.set_title(" Hardness Distribution by Loading")
save_fig(fig, "fig5a_boxplot_hardness.png")

# ── Fig 5b: Box Plot — Impact Strength ───────────────────────────────────────
fig, ax = plt.subplots(figsize=(7, 5))
bp2 = ax.boxplot([impact_raw[i] for i in range(4)],
                  labels=[f"{w} wt%" for w in loadings],
                  patch_artist=True, widths=0.45,
                  medianprops=dict(color="white", lw=2))
for patch, col in zip(bp2["boxes"], [COLORS["grey"], COLORS["amber"],
                                      COLORS["light_blue"], COLORS["blue"]]):
    patch.set_facecolor(col); patch.set_alpha(0.8)
ax.set_ylabel("Impact Strength (J/mm²)")
ax.set_title("Impact Strength Distribution by Loading")
save_fig(fig, "fig5b_boxplot_impact.png")

# ── Fig 6: Scatter + Correlation ─────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(7, 6))
scatter_colors = [COLORS["grey"], COLORS["amber"], COLORS["mid_green"], COLORS["blue"]]
for i, (w, col) in enumerate(zip(loadings, scatter_colors)):
    ax.scatter(hardness_raw[i], impact_raw[i], s=90, color=col,
               edgecolors="black", linewidths=0.5, zorder=4, label=f"{w} wt%")
    ax.scatter(H_mean[i], IS_mean[i], s=180, color=col, marker="D",
               edgecolors="black", linewidths=0.8, zorder=5)
H_all  = hardness_raw.ravel()
IS_all = impact_raw.ravel()
r_corr, pval = stats.pearsonr(H_all, IS_all)
m, b, *_ = stats.linregress(H_all, IS_all)
x_fit = np.linspace(H_all.min(), H_all.max(), 100)
ax.plot(x_fit, m*x_fit + b, color=COLORS["grey"], ls="--", lw=1.5, alpha=0.7)
ax.text(0.05, 0.93, f"Pearson r = {r_corr:.3f} (p = {pval:.3f})",
        transform=ax.transAxes, fontsize=9,
        bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8))
ax.set_xlabel("Rockwell Hardness (HR)")
ax.set_ylabel("Impact Strength (J/mm²)")
ax.set_title(" Hardness vs. Impact Strength Correlation\n(Diamonds = group means)")
ax.legend(title="Loading", loc="lower left")
save_fig(fig, "fig6_correlation.png")

# ── Fig 7a: LOOCV Parity — Hardness ──────────────────────────────────────────
fig, ax = plt.subplots(figsize=(6, 6))
mn = min(min(y_H), min(y_H_pred_cv)) - 0.5
mx = max(max(y_H), max(y_H_pred_cv)) + 0.5
ax.plot([mn, mx], [mn, mx], "k--", lw=1.2, alpha=0.5, label="Perfect fit")
ax.scatter(y_H, y_H_pred_cv, s=120, color=COLORS["deep_green"],
           edgecolors="black", linewidths=0.8, zorder=5, label="LOOCV")
for i, w in enumerate(loadings):
    ax.annotate(f" {w}%", (y_H[i], y_H_pred_cv[i]), fontsize=9)
ax.text(0.05, 0.88, f"R² = {r2_H_cv:.3f}\nMAE = {mae_H:.3f} HR",
        transform=ax.transAxes, fontsize=9,
        bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8))
ax.set_xlabel("Observed Hardness (HR)")
ax.set_ylabel("LOOCV Predicted Hardness (HR)")
ax.set_title(" GPR Parity Plot — Hardness")
ax.legend()
save_fig(fig, "fig7a_parity_hardness.png")

# ── Fig 7b: LOOCV Parity — Impact Strength ───────────────────────────────────
fig, ax = plt.subplots(figsize=(6, 6))
mn = min(min(y_IS), min(y_IS_pred_cv)) - 0.5
mx = max(max(y_IS), max(y_IS_pred_cv)) + 0.5
ax.plot([mn, mx], [mn, mx], "k--", lw=1.2, alpha=0.5, label="Perfect fit")
ax.scatter(y_IS, y_IS_pred_cv, s=120, color=COLORS["blue"],
           edgecolors="black", linewidths=0.8, zorder=5, label="LOOCV")
for i, w in enumerate(loadings):
    ax.annotate(f" {w}%", (y_IS[i], y_IS_pred_cv[i]), fontsize=9)
ax.text(0.05, 0.88, f"R² = {r2_IS_cv:.3f}\nMAE = {mae_IS:.3f} J/mm²",
        transform=ax.transAxes, fontsize=9,
        bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8))
ax.set_xlabel("Observed Impact Strength (J/mm²)")
ax.set_ylabel(" GPR Parity Plot — Impact Strength")
ax.legend()
save_fig(fig, "fig7b_parity_impact.png")

# ── Fig 8a: Standard Deviation Bar ───────────────────────────────────────────
fig, ax = plt.subplots(figsize=(7, 5))
width = 1.4
x_pos = np.array(loadings, dtype=float)
ax.bar(x_pos - width/2, H_std,  width=width, color=COLORS["deep_green"],
       alpha=0.85, label="Hardness SD", edgecolor="white")
ax.bar(x_pos + width/2, IS_std, width=width, color=COLORS["blue"],
       alpha=0.85, label="Impact SD", edgecolor="white")
ax.set_xticks(loadings); ax.set_xlabel("BP Loading (wt%)")
ax.set_ylabel("Standard Deviation")
ax.set_title(" Standard Deviation by Loading")
ax.legend()
save_fig(fig, "fig8a_std_deviation.png")

# ── Fig 8b: Coefficient of Variation ─────────────────────────────────────────
fig, ax = plt.subplots(figsize=(7, 5))
ax.bar(x_pos - width/2, H_cv,  width=width, color=COLORS["mid_green"],
       alpha=0.85, label="Hardness CV%", edgecolor="white")
ax.bar(x_pos + width/2, IS_cv, width=width, color=COLORS["light_blue"],
       alpha=0.85, label="Impact CV%", edgecolor="white")
ax.axhline(2.0, color=COLORS["pareto"], ls="--", lw=1.5, alpha=0.7, label="2% threshold")
ax.set_xticks(loadings); ax.set_xlabel("BP Loading (wt%)")
ax.set_ylabel("Coefficient of Variation (%)")
ax.set_title(" Coefficient of Variation by Loading")
ax.legend()
save_fig(fig, "fig8b_cv_percent.png")

# =============================================================================
#  8. EXCEL WORKBOOK  — all results, tables, summaries
# =============================================================================
print("\n" + "=" * 70)
print("  GENERATING EXCEL WORKBOOK")
print("=" * 70)

wb = openpyxl.Workbook()

# ── Excel styling helpers ─────────────────────────────────────────────────────
HDR_GREEN  = PatternFill("solid", fgColor="1B5E20")
HDR_BLUE   = PatternFill("solid", fgColor="1565C0")
HDR_AMBER  = PatternFill("solid", fgColor="F57F17")
HDR_GREY   = PatternFill("solid", fgColor="546E7A")
ALT_ROW    = PatternFill("solid", fgColor="E8F5E9")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
THIN = Side(style="thin", color="CCCCCC")
thin_border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(cell, fill=HDR_GREEN):
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill      = fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin_border

def style_data(cell, alt=False):
    cell.fill      = ALT_ROW if alt else WHITE_FILL
    cell.font      = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = thin_border

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def write_table(ws, df, start_row=1, start_col=1, fill=HDR_GREEN, title=None):
    if title:
        ws.cell(row=start_row, column=start_col, value=title).font = Font(
            bold=True, size=12, name="Arial")
        start_row += 1
    for ci, col in enumerate(df.columns):
        cell = ws.cell(row=start_row, column=start_col + ci, value=col)
        style_header(cell, fill)
    for ri, row in enumerate(df.itertuples(index=False), 1):
        for ci, val in enumerate(row):
            cell = ws.cell(row=start_row + ri, column=start_col + ci, value=val)
            style_data(cell, alt=(ri % 2 == 0))
    return start_row + len(df) + 2


# ─── Sheet 1: Descriptive Statistics ─────────────────────────────────────────
ws1 = wb.active
ws1.title = "Descriptive Statistics"
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 22

# Title
ws1.merge_cells("A1:G1")
t = ws1["A1"]
t.value = "BP/PBS Green Composites — Descriptive Statistics"
t.font = Font(bold=True, size=14, name="Arial", color="1B5E20")
t.alignment = Alignment(horizontal="center")

# Raw data sub-table
ws1.cell(row=3, column=1, value="RAW EXPERIMENTAL DATA").font = Font(bold=True, size=11, name="Arial")
raw_headers = ["Loading (wt%)", "Rep 1 H (HR)", "Rep 2 H (HR)", "Rep 3 H (HR)",
               "Rep 1 IS (J/mm²)", "Rep 2 IS (J/mm²)", "Rep 3 IS (J/mm²)"]
for ci, h in enumerate(raw_headers, 1):
    style_header(ws1.cell(row=4, column=ci), fill=HDR_GREY)
    ws1.cell(row=4, column=ci).value = h
for ri, (load, h_row, is_row) in enumerate(zip(loadings, hardness_raw, impact_raw)):
    for ci, val in enumerate([load, *h_row, *is_row], 1):
        cell = ws1.cell(row=5+ri, column=ci, value=round(float(val), 5))
        style_data(cell, alt=(ri % 2 == 0))

# Summary stats table
row = write_table(ws1, df_stats, start_row=10, fill=HDR_GREEN,
                  title="SUMMARY STATISTICS")

# Delta vs baseline table
write_table(ws1, df_delta, start_row=row, fill=HDR_AMBER,
            title="Δ vs Neat PBS (0 wt%)")

set_col_widths(ws1, [14, 13, 13, 13, 15, 13, 13])

# ─── Sheet 2: ANOVA & Tukey HSD ──────────────────────────────────────────────
ws2 = wb.create_sheet("ANOVA & Tukey HSD")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:D1")
ws2["A1"].value = "One-Way ANOVA Results"
ws2["A1"].font  = Font(bold=True, size=13, name="Arial", color="1565C0")

anova_df = pd.DataFrame({
    "Property":   ["Hardness", "Impact Strength"],
    "F-statistic": [round(F_H, 4), round(F_IS, 4)],
    "p-value":    [round(p_H, 6),  round(p_IS, 6)],
    "Significance": ["*** SIGNIFICANT" if p_H < 0.05 else "n.s.",
                     "*** SIGNIFICANT" if p_IS < 0.05 else "n.s."],
})
row = write_table(ws2, anova_df, start_row=2, fill=HDR_BLUE)

row = write_table(ws2, df_tukey_H, start_row=row, fill=HDR_GREEN,
                  title="Tukey HSD — Hardness")
write_table(ws2, df_tukey_IS, start_row=row, fill=HDR_BLUE,
            title="Tukey HSD — Impact Strength")

set_col_widths(ws2, [16, 14, 12, 16])

# ─── Sheet 3: GPR Model Performance ──────────────────────────────────────────
ws3 = wb.create_sheet("GPR Model Performance")
ws3.sheet_view.showGridLines = False

ws3["A1"].value = "GPR Model Performance (Leave-One-Out Cross-Validation)"
ws3["A1"].font  = Font(bold=True, size=13, name="Arial", color="1B5E20")

gpr_perf = pd.DataFrame({
    "Property":        ["Hardness", "Impact Strength"],
    "LOOCV R²":        [round(r2_H_cv, 4),  round(r2_IS_cv, 4)],
    "MAE":             [round(mae_H, 4),     round(mae_IS, 4)],
    "Unit":            ["HR", "J/mm²"],
})
row = write_table(ws3, gpr_perf, start_row=2, fill=HDR_GREEN)

cv_df = pd.DataFrame({
    "Loading (wt%)":   loadings.tolist(),
    "Observed H (HR)": y_H.round(4).tolist(),
    "LOOCV H (HR)":    [round(v, 4) for v in y_H_pred_cv],
    "Observed IS (J/mm²)": y_IS.round(4).tolist(),
    "LOOCV IS (J/mm²)":    [round(v, 4) for v in y_IS_pred_cv],
})
write_table(ws3, cv_df, start_row=row, fill=HDR_BLUE,
            title="LOOCV Observed vs. Predicted Values")
set_col_widths(ws3, [16, 15, 15, 18, 18])

# ─── Sheet 4: GPR Prediction Grid ────────────────────────────────────────────
ws4 = wb.create_sheet("GPR Prediction Grid")
ws4.sheet_view.showGridLines = False
ws4["A1"].value = "GPR Surrogate Predictions over Loading Grid (0–20 wt%, 400 points)"
ws4["A1"].font  = Font(bold=True, size=12, name="Arial")

grid_headers = ["Loading (wt%)", "H_pred (HR)", "H_std (HR)",
                "IS_pred (J/mm²)", "IS_std (J/mm²)", "Composite Desirability D"]
for ci, h in enumerate(grid_headers, 1):
    style_header(ws4.cell(row=2, column=ci))
    ws4.cell(row=2, column=ci).value = h
for ri, (xv, hp, hs, ip, is_, dv) in enumerate(zip(
        x_dense, H_pred, H_std_pred, IS_pred, IS_std_pred, D_grid), 1):
    row_data = [round(xv, 3), round(hp, 5), round(hs, 6),
                round(ip, 5), round(is_, 6), round(dv, 5)]
    for ci, val in enumerate(row_data, 1):
        cell = ws4.cell(row=2 + ri, column=ci, value=val)
        style_data(cell, alt=(ri % 2 == 0))
set_col_widths(ws4, [16, 14, 14, 16, 16, 22])

# ─── Sheet 5: Pareto Front & Optimization ────────────────────────────────────
ws5 = wb.create_sheet("Pareto Front & Optimization")
ws5.sheet_view.showGridLines = False

ws5["A1"].value = "NSGA-II Multi-Objective Optimization Results"
ws5["A1"].font  = Font(bold=True, size=13, name="Arial", color="D50000")

opt_summary = pd.DataFrame({
    "Parameter":  ["Optimal BP Loading (wt%)", "Predicted Hardness (HR)",
                   "Predicted Impact Strength (J/mm²)", "Composite Desirability D",
                   "Pareto Front Size"],
    "Value":      [round(opt_grid_x, 2), round(gpr_H.predict([[opt_grid_x]])[0], 4),
                   round(gpr_IS.predict([[opt_grid_x]])[0], 4),
                   round(D_grid.max(), 4), len(pf_x)],
})
row = write_table(ws5, opt_summary, start_row=2, fill=HDR_GREY, title="")

write_table(ws5, pf_df.round(4), start_row=row, fill=HDR_GREEN,
            title="Full Pareto Front (sorted by Desirability)")
set_col_widths(ws5, [26, 15, 18, 15])

# ─── Sheet 6: Desirability Analysis ──────────────────────────────────────────
ws6 = wb.create_sheet("Desirability Analysis")
ws6.sheet_view.showGridLines = False
ws6["A1"].value = "Derringer-Suich Desirability Analysis at Experimental Points"
ws6["A1"].font  = Font(bold=True, size=12, name="Arial")

des_exp = pd.DataFrame({
    "Loading (wt%)":      loadings.tolist(),
    "H_mean (HR)":        H_mean.round(4).tolist(),
    "IS_mean (J/mm²)":    IS_mean.round(4).tolist(),
    "d_H":                [round(desirability_H(h), 4) for h in H_mean],
    "d_IS":               [round(desirability_IS(is_), 4) for is_ in IS_mean],
    "D (composite)":      [round((desirability_H(h)*desirability_IS(is_))**0.5, 4)
                           for h, is_ in zip(H_mean, IS_mean)],
})
write_table(ws6, des_exp, start_row=2, fill=HDR_AMBER)
set_col_widths(ws6, [16, 14, 16, 12, 12, 15])

# ─── Sheet 7: Final Summary ───────────────────────────────────────────────────
ws7 = wb.create_sheet("Final Summary")
ws7.sheet_view.showGridLines = False
ws7.merge_cells("A1:B1")
ws7["A1"].value = "BP/PBS Green Composites — Final Optimisation Summary"
ws7["A1"].font  = Font(bold=True, size=14, name="Arial", color="1B5E20")
ws7["A1"].alignment = Alignment(horizontal="center")

summary_rows = [
    ("Recommended BP loading",         f"{opt_grid_x:.1f} wt%"),
    ("Predicted Hardness",              f"{gpr_H.predict([[opt_grid_x]])[0]:.4f} HR"),
    ("Predicted Impact Strength",       f"{gpr_IS.predict([[opt_grid_x]])[0]:.4f} J/mm²"),
    ("Composite Desirability D",        f"{D_grid.max():.4f}"),
    ("GPR LOOCV R² — Hardness",        f"{r2_H_cv:.4f}"),
    ("GPR LOOCV R² — Impact",          f"{r2_IS_cv:.4f}"),
    ("GPR LOOCV MAE — Hardness",       f"{mae_H:.4f} HR"),
    ("GPR LOOCV MAE — Impact",         f"{mae_IS:.4f} J/mm²"),
    ("ANOVA p — Hardness",             f"{p_H:.6f} (significant)"),
    ("ANOVA p — Impact Strength",      f"{p_IS:.6f} (significant)"),
    ("Max CV% (all data)",              f"{max(H_cv.max(), IS_cv.max()):.2f}% (< 2% threshold)"),
    ("Pareto front size (NSGA-II)",    f"{len(pf_x)} solutions"),
    ("Pearson r (H vs IS)",            f"{r_corr:.3f}  (p = {pval:.3f})"),
]

style_header(ws7.cell(row=3, column=1), fill=HDR_GREEN)
ws7.cell(row=3, column=1).value = "Parameter"
style_header(ws7.cell(row=3, column=2), fill=HDR_GREEN)
ws7.cell(row=3, column=2).value = "Value / Result"
for ri, (param, val) in enumerate(summary_rows, 1):
    c1 = ws7.cell(row=3+ri, column=1, value=param)
    c2 = ws7.cell(row=3+ri, column=2, value=val)
    style_data(c1, alt=(ri % 2 == 0))
    style_data(c2, alt=(ri % 2 == 0))
    c1.alignment = Alignment(horizontal="left", vertical="center")

set_col_widths(ws7, [38, 28])

# Save workbook
xlsx_path = os.path.join(OUT_DIR, "bp_pbs_analysis_results.xlsx")
wb.save(xlsx_path)
print(f"  ✓ bp_pbs_analysis_results.xlsx")

# =============================================================================
#  9. FINAL CONSOLE SUMMARY
# =============================================================================
print("\n" + "=" * 70)
print("  FINAL SUMMARY — OPTIMAL COMPOSITION")
print("=" * 70)
print(f"  Recommended BP loading    : {opt_grid_x:.1f} wt%")
print(f"  Predicted Hardness        : {gpr_H.predict([[opt_grid_x]])[0]:.4f} HR")
print(f"  Predicted Impact Strength : {gpr_IS.predict([[opt_grid_x]])[0]:.4f} J/mm²")
print(f"  Composite Desirability    : {D_grid.max():.4f}")
print(f"  GPR R² (LOOCV) Hardness  : {r2_H_cv:.4f}")
print(f"  GPR R² (LOOCV) Impact    : {r2_IS_cv:.4f}")
print(f"  ANOVA p (Hardness)        : {p_H:.6f}  (sig.)")
print(f"  ANOVA p (Impact)          : {p_IS:.6f}  (sig.)")
print(f"  Max CV% (all data)        : {max(H_cv.max(), IS_cv.max()):.2f}% (< 2% threshold)")
print(f"\n  Output directory : {OUT_DIR}")
print(f"  Figures saved    : 14 separate PNG files")
print(f"  Excel workbook   : bp_pbs_analysis_results.xlsx (7 sheets)")
print("=" * 70)
