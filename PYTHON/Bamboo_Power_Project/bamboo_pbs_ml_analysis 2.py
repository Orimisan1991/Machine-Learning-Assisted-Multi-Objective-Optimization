"""
=============================================================================
Machine Learning-Based Prediction of Mechanical Properties of
Bamboo Powder-Reinforced PBS Green Composites
=============================================================================
Models: Random Forest | XGBoost | Artificial Neural Network (ANN)
Targets: Tensile Strength | Flexural Strength | Hardness | Impact Strength
=============================================================================
Requirements:
    pip install numpy pandas scikit-learn xgboost tensorflow matplotlib seaborn openpyxl
=============================================================================
"""

import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.patches import FancyBboxPatch
import seaborn as sns
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import MinMaxScaler
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error
from sklearn.model_selection import LeaveOneOut, cross_val_predict
import xgboost as xgb
import tensorflow as tf
from tensorflow import keras
from tensorflow.keras import layers, callbacks
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# 0.  REPRODUCIBILITY
# ─────────────────────────────────────────────
np.random.seed(42)
tf.random.set_seed(42)

# Output folders
os.makedirs("figures", exist_ok=True)

# ─────────────────────────────────────────────
# 1.  DATASET
# ─────────────────────────────────────────────
raw_data = {
    "Bamboo_pct": [0, 0, 0, 5, 5, 5, 10, 10, 10, 20, 20, 20],
    "Tensile":    [14.19108, 13.23058, 11.14138,
                    8.70904, 11.58424, 14.94237,
                   11.82356,  9.72709, 11.96903,
                    8.43114,  5.97639, 10.17000],
    "Flexural":   [3.97348, 4.58523, 4.31535,
                   4.06532, 4.22875, 4.29750,
                   4.99372, 5.49584, 6.10767,
                   5.55638, 6.45523, 5.93439],
    "Hardness":   [65.16037, 64.74724, 64.82117,
                   64.50748, 65.53171, 66.51754,
                   63.60067, 63.31122, 63.72942,
                   62.37243, 62.47250, 63.06814],
    "Impact":     [12.39612, 12.83699, 12.86082,
                   13.81485, 14.20038, 14.65807,
                   13.99403, 14.03337, 14.14561,
                   13.44853, 13.70879, 13.69164],
}

df = pd.DataFrame(raw_data)
mean_df = df.groupby("Bamboo_pct").mean().reset_index()

TARGETS    = ["Tensile", "Flexural", "Hardness", "Impact"]
TARGET_LABELS = {
    "Tensile":  "Tensile Strength (MPa)",
    "Flexural": "Flexural Stress (MPa)",
    "Hardness": "Rockwell Hardness",
    "Impact":   "Impact Strength (J/mm²)",
}
COLORS = {
    "Random Forest": "#2ecc71",
    "XGBoost":       "#3498db",
    "ANN":           "#e74c3c",
}
PALETTE = ["#2d6a4f", "#52b788", "#b7e4c7", "#d8f3dc"]

print("=" * 65)
print("  BAMBOO POWDER / PBS GREEN COMPOSITES — ML ANALYSIS")
print("=" * 65)
print(f"\nDataset: {len(df)} observations | {len(TARGETS)} target properties")
print("\n── Mean Mechanical Properties by Loading Level ──")
print(mean_df.to_string(index=False, float_format=lambda x: f"{x:.4f}"))


# ─────────────────────────────────────────────
# 2.  PRE-PROCESSING
# ─────────────────────────────────────────────
X = df[["Bamboo_pct"]].values
Y = df[TARGETS].values

scaler_X = MinMaxScaler()
scaler_Y = MinMaxScaler()
X_scaled = scaler_X.fit_transform(X)
Y_scaled = scaler_Y.fit_transform(Y)


# ─────────────────────────────────────────────
# 3.  HELPER: METRICS
# ─────────────────────────────────────────────
def compute_metrics(y_true, y_pred):
    r2   = r2_score(y_true, y_pred)
    rmse = np.sqrt(mean_squared_error(y_true, y_pred))
    mae  = mean_absolute_error(y_true, y_pred)
    return r2, rmse, mae


# ─────────────────────────────────────────────
# 4.  RANDOM FOREST
# ─────────────────────────────────────────────
print("\n" + "─" * 65)
print("  [1/3]  RANDOM FOREST  (Leave-One-Out CV)")
print("─" * 65)

rf_model = RandomForestRegressor(
    n_estimators=300,
    max_depth=5,
    min_samples_leaf=2,
    max_features=1.0,
    random_state=42
)

loo = LeaveOneOut()
rf_preds = np.zeros_like(Y)
for train_idx, test_idx in loo.split(X_scaled):
    rf_model.fit(X_scaled[train_idx], Y_scaled[train_idx])
    pred_scaled = rf_model.predict(X_scaled[test_idx])
    rf_preds[test_idx] = scaler_Y.inverse_transform(pred_scaled.reshape(1, -1))

rf_metrics = {}
for i, t in enumerate(TARGETS):
    r2, rmse, mae = compute_metrics(Y[:, i], rf_preds[:, i])
    rf_metrics[t] = {"R2": r2, "RMSE": rmse, "MAE": mae}
    print(f"  {t:<10}  R²={r2:.4f}  RMSE={rmse:.4f}  MAE={mae:.4f}")

rf_model.fit(X_scaled, Y_scaled)
X_dense = np.linspace(0, 20, 200).reshape(-1, 1)
X_dense_scaled = scaler_X.transform(X_dense)
rf_curve = scaler_Y.inverse_transform(rf_model.predict(X_dense_scaled))


# ─────────────────────────────────────────────
# 5.  XGBOOST
# ─────────────────────────────────────────────
print("\n" + "─" * 65)
print("  [2/3]  XGBOOST  (Leave-One-Out CV)")
print("─" * 65)

xgb_models = []
xgb_preds  = np.zeros_like(Y)

for i, t in enumerate(TARGETS):
    model = xgb.XGBRegressor(
        n_estimators=300,
        max_depth=3,
        learning_rate=0.05,
        subsample=0.8,
        colsample_bytree=0.8,
        reg_alpha=0.1,
        reg_lambda=1.0,
        random_state=42,
        verbosity=0
    )
    loo_preds = []
    for train_idx, test_idx in loo.split(X_scaled):
        model.fit(X_scaled[train_idx], Y_scaled[train_idx, i])
        p = model.predict(X_scaled[test_idx])
        loo_preds.append(p[0])
    xgb_preds[:, i] = scaler_Y.inverse_transform(
        np.array(loo_preds).reshape(-1, 1) *
        (scaler_Y.data_range_[i]) + scaler_Y.data_min_[i]
        if False else
        np.column_stack([
            np.zeros((len(loo_preds), i)),
            np.array(loo_preds).reshape(-1, 1),
            np.zeros((len(loo_preds), len(TARGETS) - i - 1))
        ])
    )[:, i]

    y_min = Y[:, i].min()
    y_max = Y[:, i].max()
    xgb_preds[:, i] = np.array(loo_preds) * (y_max - y_min) + y_min

    model.fit(X_scaled, Y_scaled[:, i])
    xgb_models.append(model)

    r2, rmse, mae = compute_metrics(Y[:, i], xgb_preds[:, i])
    print(f"  {t:<10}  R²={r2:.4f}  RMSE={rmse:.4f}  MAE={mae:.4f}")

xgb_metrics = {}
for i, t in enumerate(TARGETS):
    r2, rmse, mae = compute_metrics(Y[:, i], xgb_preds[:, i])
    xgb_metrics[t] = {"R2": r2, "RMSE": rmse, "MAE": mae}

xgb_curve = np.zeros((200, len(TARGETS)))
for i, model in enumerate(xgb_models):
    raw = model.predict(X_dense_scaled)
    y_min = Y[:, i].min(); y_max = Y[:, i].max()
    xgb_curve[:, i] = raw * (y_max - y_min) + y_min


# ─────────────────────────────────────────────
# 6.  ARTIFICIAL NEURAL NETWORK
# ─────────────────────────────────────────────
print("\n" + "─" * 65)
print("  [3/3]  ANN  (Leave-One-Out CV)")
print("─" * 65)

def build_ann(input_dim=1, output_dim=4):
    inp = keras.Input(shape=(input_dim,))
    x = layers.Dense(64, activation="relu")(inp)
    x = layers.BatchNormalization()(x)
    x = layers.Dense(32, activation="relu")(x)
    x = layers.BatchNormalization()(x)
    x = layers.Dense(16, activation="relu")(x)
    x = layers.BatchNormalization()(x)
    out = layers.Dense(output_dim, activation="linear")(x)
    model = keras.Model(inp, out)
    model.compile(optimizer=keras.optimizers.Adam(1e-3), loss="mse")
    return model

ann_preds = np.zeros_like(Y_scaled)
for train_idx, test_idx in loo.split(X_scaled):
    ann = build_ann()
    es  = callbacks.EarlyStopping(monitor="val_loss", patience=80,
                                   restore_best_weights=True, verbose=0)
    ann.fit(
        X_scaled[train_idx], Y_scaled[train_idx],
        validation_split=0.2,
        epochs=1000, batch_size=4, verbose=0,
        callbacks=[es]
    )
    ann_preds[test_idx] = ann.predict(X_scaled[test_idx], verbose=0)

ann_preds_orig = scaler_Y.inverse_transform(ann_preds)

ann_metrics = {}
for i, t in enumerate(TARGETS):
    r2, rmse, mae = compute_metrics(Y[:, i], ann_preds_orig[:, i])
    ann_metrics[t] = {"R2": r2, "RMSE": rmse, "MAE": mae}
    print(f"  {t:<10}  R²={r2:.4f}  RMSE={rmse:.4f}  MAE={mae:.4f}")

ann_full = build_ann()
es_full = callbacks.EarlyStopping(monitor="val_loss", patience=100,
                                   restore_best_weights=True, verbose=0)
ann_full.fit(X_scaled, Y_scaled, validation_split=0.2,
             epochs=1500, batch_size=4, verbose=0, callbacks=[es_full])
ann_curve = scaler_Y.inverse_transform(ann_full.predict(X_dense_scaled, verbose=0))


# ─────────────────────────────────────────────
# 7.  METRICS SUMMARY TABLE
# ─────────────────────────────────────────────
print("\n" + "=" * 65)
print("  PERFORMANCE METRICS SUMMARY")
print("=" * 65)

records = []
for model_name, metrics in [("Random Forest", rf_metrics),
                             ("XGBoost",       xgb_metrics),
                             ("ANN",           ann_metrics)]:
    for t in TARGETS:
        records.append({
            "Model": model_name,
            "Property": t,
            "R²":    round(metrics[t]["R2"],   4),
            "RMSE":  round(metrics[t]["RMSE"], 4),
            "MAE":   round(metrics[t]["MAE"],  4),
        })

metrics_df = pd.DataFrame(records)
print(metrics_df.to_string(index=False))


# ═══════════════════════════════════════════════════════
#  EXCEL EXPORT
# ═══════════════════════════════════════════════════════

def style_header(cell, bg_color="2d6a4f", font_color="FFFFFF"):
    cell.font = Font(bold=True, color=font_color, name="Arial", size=11)
    cell.fill = PatternFill("solid", start_color=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def style_subheader(cell, bg_color="52b788"):
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def style_data(cell, bold=False):
    cell.font = Font(name="Arial", size=10, bold=bold)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def thin_border():
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

print("\n" + "─" * 65)
print("  EXPORTING RESULTS TO EXCEL")
print("─" * 65)

wb = Workbook()

# ── Sheet 1: Raw Dataset ──────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Raw Dataset"

ws1.merge_cells("A1:F1")
ws1["A1"] = "Bamboo Powder PBS Composites — Raw Experimental Data"
style_header(ws1["A1"])
ws1.row_dimensions[1].height = 22

headers = ["Sample No.", "Bamboo Content (%)", "Tensile Strength (MPa)",
           "Flexural Stress (MPa)", "Rockwell Hardness", "Impact Strength (J/mm²)"]
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=2, column=col, value=h)
    style_subheader(cell)

for row_i, (_, row) in enumerate(df.iterrows(), 3):
    ws1.cell(row=row_i, column=1, value=row_i - 2)
    ws1.cell(row=row_i, column=2, value=row["Bamboo_pct"])
    ws1.cell(row=row_i, column=3, value=round(row["Tensile"],  5))
    ws1.cell(row=row_i, column=4, value=round(row["Flexural"], 5))
    ws1.cell(row=row_i, column=5, value=round(row["Hardness"], 5))
    ws1.cell(row=row_i, column=6, value=round(row["Impact"],   5))
    for col in range(1, 7):
        cell = ws1.cell(row=row_i, column=col)
        style_data(cell)
        cell.border = thin_border()
        if row_i % 2 == 0:
            cell.fill = PatternFill("solid", start_color="f0f9f4")

auto_width(ws1)


# ── Sheet 2: Mean Summary ─────────────────────────────────────────────────────
ws2 = wb.create_sheet("Mean Summary")

ws2.merge_cells("A1:F1")
ws2["A1"] = "Mean Mechanical Properties by Bamboo Loading Level"
style_header(ws2["A1"])
ws2.row_dimensions[1].height = 22

headers2 = ["Bamboo Content (%)", "Mean Tensile (MPa)", "Mean Flexural (MPa)",
            "Mean Hardness", "Mean Impact (J/mm²)", "Std Tensile", "Std Flexural",
            "Std Hardness", "Std Impact"]
ws2.merge_cells("A1:I1")
ws2["A1"] = "Mean Mechanical Properties by Bamboo Loading Level"
style_header(ws2["A1"])
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=2, column=col, value=h)
    style_subheader(cell)

std_df = df.groupby("Bamboo_pct").std(ddof=1).reset_index()
for row_i, pct in enumerate([0, 5, 10, 20], 3):
    m = mean_df[mean_df.Bamboo_pct == pct].iloc[0]
    s = std_df[std_df.Bamboo_pct == pct].iloc[0]
    vals = [pct, round(m.Tensile, 4), round(m.Flexural, 4),
            round(m.Hardness, 4), round(m.Impact, 4),
            round(s.Tensile, 4), round(s.Flexural, 4),
            round(s.Hardness, 4), round(s.Impact, 4)]
    for col, v in enumerate(vals, 1):
        cell = ws2.cell(row=row_i, column=col, value=v)
        style_data(cell)
        cell.border = thin_border()

auto_width(ws2)


# ── Sheet 3: Model Metrics ────────────────────────────────────────────────────
ws3 = wb.create_sheet("Model Metrics")

ws3.merge_cells("A1:E1")
ws3["A1"] = "ML Model Performance Metrics (Leave-One-Out Cross-Validation)"
style_header(ws3["A1"])
ws3.row_dimensions[1].height = 22

for col, h in enumerate(["Model", "Property", "R²", "RMSE", "MAE"], 1):
    cell = ws3.cell(row=2, column=col, value=h)
    style_subheader(cell)

model_colors = {"Random Forest": "d5f5e3", "XGBoost": "d6eaf8", "ANN": "fce4e4"}
for row_i, rec in enumerate(records, 3):
    bg = model_colors.get(rec["Model"], "FFFFFF")
    for col, val in enumerate([rec["Model"], rec["Property"],
                                rec["R²"], rec["RMSE"], rec["MAE"]], 1):
        cell = ws3.cell(row=row_i, column=col, value=val)
        style_data(cell, bold=(col == 1))
        cell.fill = PatternFill("solid", start_color=bg)
        cell.border = thin_border()

# Best model per property block
ws3.cell(row=len(records) + 4, column=1, value="Best Model per Property (by R²)")
style_subheader(ws3.cell(row=len(records) + 4, column=1))
ws3.merge_cells(f"A{len(records)+4}:E{len(records)+4}")

for col, h in enumerate(["Property", "Best Model", "R²", "RMSE", "MAE"], 1):
    cell = ws3.cell(row=len(records) + 5, column=col, value=h)
    style_subheader(cell, bg_color="457b9d")

for r_i, t in enumerate(TARGETS, len(records) + 6):
    scores = {
        "Random Forest": rf_metrics[t]["R2"],
        "XGBoost":       xgb_metrics[t]["R2"],
        "ANN":           ann_metrics[t]["R2"],
    }
    best = max(scores, key=scores.get)
    bm = {"Random Forest": rf_metrics, "XGBoost": xgb_metrics, "ANN": ann_metrics}[best]
    row_vals = [t, best, round(bm[t]["R2"], 4), round(bm[t]["RMSE"], 4), round(bm[t]["MAE"], 4)]
    for col, v in enumerate(row_vals, 1):
        cell = ws3.cell(row=r_i, column=col, value=v)
        style_data(cell, bold=(col <= 2))
        cell.fill = PatternFill("solid", start_color="fff3cd")
        cell.border = thin_border()

auto_width(ws3)


# ── Sheet 4: LOOCV Predictions ───────────────────────────────────────────────
ws4 = wb.create_sheet("LOOCV Predictions")

ws4.merge_cells("A1:M1")
ws4["A1"] = "Leave-One-Out CV Predictions vs Actual Values"
style_header(ws4["A1"])
ws4.row_dimensions[1].height = 22

pred_headers = (["Sample", "Bamboo (%)"] +
                [f"Actual {t}" for t in TARGETS] +
                [f"RF {t}" for t in TARGETS] +
                [f"XGB {t}" for t in TARGETS] +
                [f"ANN {t}" for t in TARGETS])
for col, h in enumerate(pred_headers, 1):
    cell = ws4.cell(row=2, column=col, value=h)
    style_subheader(cell)

for row_i in range(len(df)):
    row_data = (
        [row_i + 1, df["Bamboo_pct"].iloc[row_i]] +
        [round(Y[row_i, j], 5)              for j in range(4)] +
        [round(rf_preds[row_i, j], 5)       for j in range(4)] +
        [round(xgb_preds[row_i, j], 5)      for j in range(4)] +
        [round(ann_preds_orig[row_i, j], 5) for j in range(4)]
    )
    for col, v in enumerate(row_data, 1):
        cell = ws4.cell(row=row_i + 3, column=col, value=v)
        style_data(cell)
        cell.border = thin_border()
        if row_i % 2 == 0:
            cell.fill = PatternFill("solid", start_color="f8f9fa")

auto_width(ws4)


# ── Sheet 5: Correlation Matrix ───────────────────────────────────────────────
ws5 = wb.create_sheet("Correlation Matrix")

ws5.merge_cells("A1:F1")
ws5["A1"] = "Pearson Correlation Matrix"
style_header(ws5["A1"])
ws5.row_dimensions[1].height = 22

corr = df[["Bamboo_pct"] + TARGETS].corr()
corr_labels = ["Bamboo (%)", "Tensile", "Flexural", "Hardness", "Impact"]

for col, label in enumerate(corr_labels, 2):
    cell = ws5.cell(row=2, column=col, value=label)
    style_subheader(cell)
for row_i, label in enumerate(corr_labels, 3):
    cell = ws5.cell(row=row_i, column=1, value=label)
    style_subheader(cell)

for r, row_label in enumerate(corr.index):
    for c, col_label in enumerate(corr.columns):
        val = round(corr.loc[row_label, col_label], 4)
        cell = ws5.cell(row=r + 3, column=c + 2, value=val)
        style_data(cell)
        cell.border = thin_border()
        # Colour-code: green = positive, red = negative
        if val >= 0.7:
            cell.fill = PatternFill("solid", start_color="b7e4c7")
        elif val <= -0.7:
            cell.fill = PatternFill("solid", start_color="f5c6cb")
        elif abs(val) == 1.0:
            cell.fill = PatternFill("solid", start_color="2d6a4f")

auto_width(ws5)


# ── Sheet 6: Figure Index ─────────────────────────────────────────────────────
ws6 = wb.create_sheet("Figure Index")

ws6.merge_cells("A1:C1")
ws6["A1"] = "Generated Figures — Index"
style_header(ws6["A1"])

for col, h in enumerate(["Figure File", "Description"], 1):
    cell = ws6.cell(row=2, column=col, value=h)
    style_subheader(cell)

figure_index = [
    ("01a_experimental_Tensile.png",   "Experimental bar chart — Tensile Strength"),
    ("01b_experimental_Flexural.png",  "Experimental bar chart — Flexural Stress"),
    ("01c_experimental_Hardness.png",  "Experimental bar chart — Hardness"),
    ("01d_experimental_Impact.png",    "Experimental bar chart — Impact Strength"),
    ("02a_prediction_curves_Tensile.png",   "ML prediction curves — Tensile"),
    ("02b_prediction_curves_Flexural.png",  "ML prediction curves — Flexural"),
    ("02c_prediction_curves_Hardness.png",  "ML prediction curves — Hardness"),
    ("02d_prediction_curves_Impact.png",    "ML prediction curves — Impact"),
    ("03a_actual_vs_pred_RF_Tensile.png",   "Actual vs Predicted — RF Tensile"),
    ("03b_actual_vs_pred_RF_Flexural.png",  "Actual vs Predicted — RF Flexural"),
    ("03c_actual_vs_pred_RF_Hardness.png",  "Actual vs Predicted — RF Hardness"),
    ("03d_actual_vs_pred_RF_Impact.png",    "Actual vs Predicted — RF Impact"),
    ("03e_actual_vs_pred_XGB_Tensile.png",  "Actual vs Predicted — XGBoost Tensile"),
    ("03f_actual_vs_pred_XGB_Flexural.png", "Actual vs Predicted — XGBoost Flexural"),
    ("03g_actual_vs_pred_XGB_Hardness.png", "Actual vs Predicted — XGBoost Hardness"),
    ("03h_actual_vs_pred_XGB_Impact.png",   "Actual vs Predicted — XGBoost Impact"),
    ("03i_actual_vs_pred_ANN_Tensile.png",  "Actual vs Predicted — ANN Tensile"),
    ("03j_actual_vs_pred_ANN_Flexural.png", "Actual vs Predicted — ANN Flexural"),
    ("03k_actual_vs_pred_ANN_Hardness.png", "Actual vs Predicted — ANN Hardness"),
    ("03l_actual_vs_pred_ANN_Impact.png",   "Actual vs Predicted — ANN Impact"),
    ("04_r2_comparison.png",             "R² grouped bar chart — all models"),
    ("05a_rmse_comparison.png",          "RMSE grouped bar chart — all models"),
    ("05b_mae_comparison.png",           "MAE grouped bar chart — all models"),
    ("06_correlation_heatmap.png",       "Pearson Correlation Heatmap"),
    ("07a_residuals_RF_Tensile.png",     "Residuals — RF Tensile"),
    ("07b_residuals_RF_Flexural.png",    "Residuals — RF Flexural"),
    ("07c_residuals_RF_Hardness.png",    "Residuals — RF Hardness"),
    ("07d_residuals_RF_Impact.png",      "Residuals — RF Impact"),
    ("07e_residuals_XGB_Tensile.png",    "Residuals — XGBoost Tensile"),
    ("07f_residuals_XGB_Flexural.png",   "Residuals — XGBoost Flexural"),
    ("07g_residuals_XGB_Hardness.png",   "Residuals — XGBoost Hardness"),
    ("07h_residuals_XGB_Impact.png",     "Residuals — XGBoost Impact"),
    ("07i_residuals_ANN_Tensile.png",    "Residuals — ANN Tensile"),
    ("07j_residuals_ANN_Flexural.png",   "Residuals — ANN Flexural"),
    ("07k_residuals_ANN_Hardness.png",   "Residuals — ANN Hardness"),
    ("07l_residuals_ANN_Impact.png",     "Residuals — ANN Impact"),
    ("08_feature_importance.png",        "RF Feature Importance"),
    ("09_radar_r2.png",                  "Radar Chart — R² per model × property"),
    ("10_pairplot.png",                  "Pairplot coloured by bamboo loading"),
    ("11a_metrics_heatmap_R2.png",       "Performance Metrics Heatmap — R²"),
    ("11b_metrics_heatmap_RMSE.png",     "Performance Metrics Heatmap — RMSE"),
    ("11c_metrics_heatmap_MAE.png",      "Performance Metrics Heatmap — MAE"),
    ("12a_trend_Tensile_Flexural.png",   "Combined property trend — Tensile & Flexural"),
    ("12b_trend_Hardness_Impact.png",    "Combined property trend — Hardness & Impact"),
]

for row_i, (fname, desc) in enumerate(figure_index, 3):
    ws6.cell(row=row_i, column=1, value=fname).border = thin_border()
    ws6.cell(row=row_i, column=2, value=desc).border = thin_border()
    for col in [1, 2]:
        style_data(ws6.cell(row=row_i, column=col))
        if row_i % 2 == 0:
            ws6.cell(row=row_i, column=col).fill = PatternFill("solid", start_color="f0f9f4")

auto_width(ws6)
# ============================================================
# Sheet: Prediction Curves
# ============================================================

ws7 = wb.create_sheet("Prediction Curves")

headers = [
    "Bamboo %",
    "RF Tensile","RF Flexural","RF Hardness","RF Impact",
    "XGB Tensile","XGB Flexural","XGB Hardness","XGB Impact",
    "ANN Tensile","ANN Flexural","ANN Hardness","ANN Impact"
]

for c,h in enumerate(headers,1):
    cell = ws7.cell(row=1,column=c,value=h)
    style_subheader(cell)

for i in range(len(X_dense)):
    values = [
        float(X_dense[i]),
        *rf_curve[i],
        *xgb_curve[i],
        *ann_curve[i]
    ]

    for c,v in enumerate(values,1):
        cell=ws7.cell(row=i+2,column=c,value=float(v))
        style_data(cell)
        cell.border=thin_border()

auto_width(ws7)
# ============================================================
# Sheet: Residuals
# ============================================================

ws8 = wb.create_sheet("Residuals")

headers=["Sample","Property",
         "RF Residual",
         "XGB Residual",
         "ANN Residual"]

for c,h in enumerate(headers,1):
    style_subheader(ws8.cell(row=1,column=c,value=h))

row=2

for j,target in enumerate(TARGETS):

    for i in range(len(df)):

        ws8.cell(row=row,column=1).value=i+1
        ws8.cell(row=row,column=2).value=target
        ws8.cell(row=row,column=3).value=float(Y[i,j]-rf_preds[i,j])
        ws8.cell(row=row,column=4).value=float(Y[i,j]-xgb_preds[i,j])
        ws8.cell(row=row,column=5).value=float(Y[i,j]-ann_preds_orig[i,j])

        for c in range(1,6):
            style_data(ws8.cell(row=row,column=c))
            ws8.cell(row=row,column=c).border=thin_border()

        row+=1

auto_width(ws8)

# ============================================================
# Sheet: Feature Importance
# ============================================================

ws9=wb.create_sheet("Feature Importance")

style_subheader(ws9["A1"])
style_subheader(ws9["B1"])

ws9["A1"]="Property"
ws9["B1"]="Importance"

for i,target in enumerate(TARGETS):

    rf=RandomForestRegressor(
        n_estimators=300,
        random_state=42
    )

    rf.fit(X_scaled,Y_scaled[:,i])

    ws9.cell(row=i+2,column=1).value=target
    ws9.cell(row=i+2,column=2).value=float(rf.feature_importances_[0])

    for c in range(1,3):
        style_data(ws9.cell(row=i+2,column=c))
        ws9.cell(row=i+2,column=c).border=thin_border()

auto_width(ws9)

# ============================================================
# Sheet: Best Models
# ============================================================

ws10=wb.create_sheet("Best Models")

headers=["Property","Best Model","R²","RMSE","MAE"]

for c,h in enumerate(headers,1):
    style_subheader(ws10.cell(row=1,column=c,value=h))

row=2

for t in TARGETS:

    models={
        "Random Forest":rf_metrics[t],
        "XGBoost":xgb_metrics[t],
        "ANN":ann_metrics[t]
    }

    best=max(models,key=lambda k:models[k]["R2"])

    ws10.cell(row=row,column=1).value=t
    ws10.cell(row=row,column=2).value=best
    ws10.cell(row=row,column=3).value=models[best]["R2"]
    ws10.cell(row=row,column=4).value=models[best]["RMSE"]
    ws10.cell(row=row,column=5).value=models[best]["MAE"]

    for c in range(1,6):
        style_data(ws10.cell(row=row,column=c))
        ws10.cell(row=row,column=c).border=thin_border()

    row+=1

auto_width(ws10)
# ============================================================
# Sheet: Descriptive Statistics
# ============================================================

ws11=wb.create_sheet("Statistics")

stats=df.describe().T

for c,col in enumerate(["Variable"]+list(stats.columns),1):
    style_subheader(ws11.cell(row=1,column=c,value=col))

for r,(idx,row) in enumerate(stats.iterrows(),2):

    ws11.cell(row=r,column=1).value=idx

    for c,val in enumerate(row,2):
        ws11.cell(row=r,column=c).value=float(val)

    for cc in range(1,len(stats.columns)+2):
        style_data(ws11.cell(row=r,column=cc))
        ws11.cell(row=r,column=cc).border=thin_border()

auto_width(ws11)

excel_file = os.path.join("figures", "bamboo_pbs_ml_results.xlsx")
wb.save(excel_file)

print("="*60)
print("Excel workbook successfully created.")
print(excel_file)
print("="*60)
print("  ✔ Saved: figures/bamboo_pbs_ml_results.xlsx")


# ═══════════════════════════════════════════════════════
#  PLOTS  (each graph saved as a SEPARATE file)
# ═══════════════════════════════════════════════════════

plt.rcParams.update({
    "font.family":       "DejaVu Sans",
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "axes.grid":         True,
    "grid.alpha":        0.3,
    "grid.linestyle":    "--",
    "figure.dpi":        150,
})

LOADINGS = [0, 5, 10, 20]


# ─────────────────────────────────────────────────────────
# PLOT 1 — Experimental mechanical properties (one per target)
# ─────────────────────────────────────────────────────────
suffix = {"Tensile": "a", "Flexural": "b", "Hardness": "c", "Impact": "d"}

for t in TARGETS:
    fig, ax = plt.subplots(figsize=(7, 5))
    vals  = [df[df.Bamboo_pct == p][t].values for p in LOADINGS]
    means = [v.mean() for v in vals]
    stds  = [v.std(ddof=1) for v in vals]
    bars  = ax.bar(LOADINGS, means, width=3.5, color=PALETTE,
                   edgecolor="white", linewidth=1.2, zorder=3)
    ax.errorbar(LOADINGS, means, yerr=stds, fmt="none",
                color="#333", capsize=5, linewidth=1.5, zorder=4)
    for xi, pct in enumerate(LOADINGS):
        ys = df[df.Bamboo_pct == pct][t].values
        ax.scatter([pct]*3, ys, color="white", edgecolor="#333", s=40, zorder=5)
    ax.set_xlabel("Bamboo Powder Content (%)", fontsize=11)
    ax.set_ylabel(TARGET_LABELS[t], fontsize=11)
    ax.set_title(f"Experimental Mechanical Properties — {t}", fontsize=12, fontweight="bold")
    ax.set_xticks(LOADINGS)
    for bar, val, std in zip(bars, means, stds):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + std * 1.1,
                f"{val:.2f}", ha="center", va="bottom", fontsize=9, fontweight="bold")
    plt.tight_layout()
    fname = f"figures/01{suffix[t]}_experimental_{t}.png"
    plt.savefig(fname, bbox_inches="tight")
    plt.close()
    print(f"  ✔ Saved: {fname}")


# ─────────────────────────────────────────────────────────
# PLOT 2 — Prediction curves (one per target)
# ─────────────────────────────────────────────────────────
for i, t in enumerate(TARGETS):
    fig, ax = plt.subplots(figsize=(7, 5))
    ax.scatter(df["Bamboo_pct"], df[t], color="#333", s=55,
               zorder=5, label="Experimental", marker="D", edgecolors="white")
    ax.scatter(mean_df["Bamboo_pct"], mean_df[t], color="#333",
               s=120, zorder=6, marker="*", label="Mean")
    ax.plot(X_dense.flatten(), rf_curve[:, i],  color=COLORS["Random Forest"],
            lw=2.2, linestyle="-",  label="Random Forest")
    ax.plot(X_dense.flatten(), xgb_curve[:, i], color=COLORS["XGBoost"],
            lw=2.2, linestyle="--", label="XGBoost")
    ax.plot(X_dense.flatten(), ann_curve[:, i], color=COLORS["ANN"],
            lw=2.2, linestyle="-.", label="ANN")
    ax.set_xlabel("Bamboo Powder Content (%)", fontsize=11)
    ax.set_ylabel(TARGET_LABELS[t], fontsize=11)
    ax.set_title(f"ML Prediction Curves — {t}", fontsize=12, fontweight="bold")
    ax.set_xticks(LOADINGS)
    ax.legend(fontsize=9, framealpha=0.7)
    plt.tight_layout()
    fname = f"figures/02{suffix[t]}_prediction_curves_{t}.png"
    plt.savefig(fname, bbox_inches="tight")
    plt.close()
    print(f"  ✔ Saved: {fname}")


# ─────────────────────────────────────────────────────────
# PLOT 3 — Actual vs Predicted (one per model × target)
# ─────────────────────────────────────────────────────────
model_data = [
    ("RF",  "Random Forest", rf_preds,        rf_metrics),
    ("XGB", "XGBoost",       xgb_preds,       xgb_metrics),
    ("ANN", "ANN",           ann_preds_orig,  ann_metrics),
]
sub_letters = list("abcdefghijkl")
letter_idx = 0

for short, mname, preds, mets in model_data:
    for t in TARGETS:
        col = TARGETS.index(t)
        fig, ax = plt.subplots(figsize=(6, 5))
        y_true = Y[:, col]
        y_pred = preds[:, col]
        ax.scatter(y_true, y_pred, color=COLORS[mname],
                   edgecolors="white", s=70, zorder=3)
        lo = min(y_true.min(), y_pred.min()) * 0.97
        hi = max(y_true.max(), y_pred.max()) * 1.03
        ax.plot([lo, hi], [lo, hi], "k--", lw=1.2, label="Ideal")
        ax.set_xlim(lo, hi); ax.set_ylim(lo, hi)
        r2 = mets[t]["R2"]
        ax.text(0.05, 0.93, f"R²={r2:.4f}", transform=ax.transAxes,
                fontsize=10, fontweight="bold", color=COLORS[mname])
        ax.set_title(f"Actual vs Predicted — {mname} | {t}", fontsize=11, fontweight="bold")
        ax.set_xlabel("Actual", fontsize=11)
        ax.set_ylabel("Predicted", fontsize=11)
        ax.legend(fontsize=9)
        plt.tight_layout()
        letter = sub_letters[letter_idx]
        fname = f"figures/03{letter}_actual_vs_pred_{short}_{t}.png"
        plt.savefig(fname, bbox_inches="tight")
        plt.close()
        print(f"  ✔ Saved: {fname}")
        letter_idx += 1


# ─────────────────────────────────────────────────────────
# PLOT 4 — R² Grouped Bar Chart (single figure)
# ─────────────────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(11, 6))
x = np.arange(len(TARGETS))
w = 0.25
for j, (mname, mets) in enumerate([("Random Forest", rf_metrics),
                                    ("XGBoost",       xgb_metrics),
                                    ("ANN",           ann_metrics)]):
    r2_vals = [mets[t]["R2"] for t in TARGETS]
    bars = ax.bar(x + (j-1)*w, r2_vals, width=w, label=mname,
                  color=COLORS[mname], edgecolor="white", linewidth=1.2)
    for bar, val in zip(bars, r2_vals):
        ax.text(bar.get_x() + bar.get_width()/2,
                bar.get_height() + 0.002, f"{val:.4f}",
                ha="center", va="bottom", fontsize=7.5, rotation=90)
ax.set_xticks(x)
ax.set_xticklabels([TARGET_LABELS[t] for t in TARGETS], fontsize=10)
ax.set_ylabel("R² Score", fontsize=11)
ax.set_title("Model Comparison — R² by Mechanical Property", fontsize=13, fontweight="bold")
ax.set_ylim(0.92, 1.005)
ax.legend(fontsize=10)
ax.axhline(1.0, color="gray", lw=0.8, linestyle=":")
plt.tight_layout()
plt.savefig("figures/04_r2_comparison.png", bbox_inches="tight")
plt.close()
print("  ✔ Saved: figures/04_r2_comparison.png")


# ─────────────────────────────────────────────────────────
# PLOT 5 — RMSE and MAE as SEPARATE figures
# ─────────────────────────────────────────────────────────
for metric_key, metric_label, fig_suffix in [("RMSE", "RMSE", "a"), ("MAE", "MAE", "b")]:
    fig, ax = plt.subplots(figsize=(10, 6))
    for j, (mname, mets) in enumerate([("Random Forest", rf_metrics),
                                        ("XGBoost",       xgb_metrics),
                                        ("ANN",           ann_metrics)]):
        vals = [mets[t][metric_key] for t in TARGETS]
        bars = ax.bar(x + (j-1)*w, vals, width=w, label=mname,
                      color=COLORS[mname], edgecolor="white")
        for bar, v in zip(bars, vals):
            ax.text(bar.get_x() + bar.get_width()/2,
                    bar.get_height() + max(vals)*0.01,
                    f"{v:.4f}", ha="center", va="bottom", fontsize=7.5, rotation=90)
    ax.set_xticks(x)
    ax.set_xticklabels(TARGETS, fontsize=11)
    ax.set_ylabel(metric_label, fontsize=11)
    ax.set_title(f"Model Comparison — {metric_label}", fontsize=12, fontweight="bold")
    ax.legend(fontsize=10)
    plt.tight_layout()
    fname = f"figures/05{fig_suffix}_{metric_key.lower()}_comparison.png"
    plt.savefig(fname, bbox_inches="tight")
    plt.close()
    print(f"  ✔ Saved: {fname}")


# ─────────────────────────────────────────────────────────
# PLOT 6 — Heatmap: Pearson Correlation (single figure)
# ─────────────────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(8, 6))
corr = df[["Bamboo_pct"] + TARGETS].corr()
sns.heatmap(corr, annot=True, fmt=".3f", cmap="RdYlGn",
            linewidths=0.6, ax=ax, vmin=-1, vmax=1,
            annot_kws={"size": 11, "weight": "bold"})
ax.set_title("Pearson Correlation Matrix", fontsize=13, fontweight="bold")
plt.tight_layout()
plt.savefig("figures/06_correlation_heatmap.png", bbox_inches="tight")
plt.close()
print("  ✔ Saved: figures/06_correlation_heatmap.png")


# ─────────────────────────────────────────────────────────
# PLOT 7 — Residuals (one per model × target)
# ─────────────────────────────────────────────────────────
res_letters = list("abcdefghijkl")
res_idx = 0

for short, mname, preds in [("RF",  "Random Forest", rf_preds),
                              ("XGB", "XGBoost",       xgb_preds),
                              ("ANN", "ANN",           ann_preds_orig)]:
    for col, t in enumerate(TARGETS):
        fig, ax = plt.subplots(figsize=(7, 5))
        residuals = Y[:, col] - preds[:, col]
        ax.stem(range(len(residuals)), residuals,
                linefmt=COLORS[mname], markerfmt="o", basefmt="k-")
        ax.axhline(0, color="black", lw=1.0)
        ax.set_xlabel("Sample Index", fontsize=11)
        ax.set_ylabel("Residual (Actual − Predicted)", fontsize=11)
        ax.set_title(f"Residuals — {mname} | {t}", fontsize=12, fontweight="bold")
        rmse = np.sqrt(mean_squared_error(Y[:, col], preds[:, col]))
        ax.text(0.02, 0.93, f"RMSE={rmse:.4f}",
                transform=ax.transAxes, fontsize=10, color=COLORS[mname], fontweight="bold")
        plt.tight_layout()
        letter = res_letters[res_idx]
        fname = f"figures/07{letter}_residuals_{short}_{t}.png"
        plt.savefig(fname, bbox_inches="tight")
        plt.close()
        print(f"  ✔ Saved: {fname}")
        res_idx += 1


# ─────────────────────────────────────────────────────────
# PLOT 8 — RF Feature Importance (single figure)
# ─────────────────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(9, 5))
rf_model.fit(X_scaled, Y_scaled)
importances = []
for i, t in enumerate(TARGETS):
    rf_single = RandomForestRegressor(n_estimators=300, max_depth=5, random_state=42)
    rf_single.fit(X_scaled, Y_scaled[:, i])
    importances.append(rf_single.feature_importances_[0])

bars = ax.barh(TARGETS, importances, color=[COLORS["Random Forest"]]*4,
               edgecolor="white", height=0.5)
ax.set_xlabel("Feature Importance (Mean Impurity Decrease)", fontsize=11)
ax.set_title("Random Forest — Feature Importance\n(Bamboo Content → Each Property)",
             fontsize=12, fontweight="bold")
ax.set_xlim(0, 1.1)
for bar, v in zip(bars, importances):
    ax.text(v + 0.02, bar.get_y() + bar.get_height()/2,
            f"{v:.4f}", va="center", fontsize=10, fontweight="bold")
plt.tight_layout()
plt.savefig("figures/08_feature_importance.png", bbox_inches="tight")
plt.close()
print("  ✔ Saved: figures/08_feature_importance.png")


# ─────────────────────────────────────────────────────────
# PLOT 9 — Radar Chart (single figure)
# ─────────────────────────────────────────────────────────
N = len(TARGETS)
angles = np.linspace(0, 2 * np.pi, N, endpoint=False).tolist()
angles += angles[:1]

fig, ax = plt.subplots(figsize=(8, 8), subplot_kw=dict(polar=True))
ax.set_facecolor("#f8f9fa")
fig.patch.set_facecolor("#f8f9fa")

for mname, mets in [("Random Forest", rf_metrics),
                     ("XGBoost",       xgb_metrics),
                     ("ANN",           ann_metrics)]:
    values = [mets[t]["R2"] for t in TARGETS]
    values += values[:1]
    ax.plot(angles, values, color=COLORS[mname], lw=2.5, label=mname)
    ax.fill(angles, values, color=COLORS[mname], alpha=0.12)

ax.set_xticks(angles[:-1])
ax.set_xticklabels(TARGETS, fontsize=12, fontweight="bold")
ax.set_ylim(0.90, 1.005)
ax.set_yticks([0.92, 0.94, 0.96, 0.98, 1.00])
ax.set_yticklabels(["0.92", "0.94", "0.96", "0.98", "1.00"], fontsize=8)
ax.set_title("R² Radar Chart — Model vs Property", fontsize=13, fontweight="bold", pad=20)
ax.legend(loc="upper right", bbox_to_anchor=(1.3, 1.15), fontsize=10)
plt.tight_layout()
plt.savefig("figures/09_radar_r2.png", bbox_inches="tight")
plt.close()
print("  ✔ Saved: figures/09_radar_r2.png")


# ─────────────────────────────────────────────────────────
# PLOT 10 — Pairplot (single figure)
# ─────────────────────────────────────────────────────────
df_plot = df.copy()
df_plot["Loading"] = df_plot["Bamboo_pct"].astype(str) + " wt%"
g = sns.pairplot(df_plot[TARGETS + ["Loading"]], hue="Loading",
                 plot_kws={"s": 70, "edgecolor": "white"},
                 palette=["#1b4332", "#52b788", "#95d5b2", "#d8f3dc"],
                 diag_kind="kde", corner=False)
g.figure.suptitle("Pairplot of Mechanical Properties by Bamboo Loading",
                  y=1.02, fontsize=13, fontweight="bold")
plt.savefig("figures/10_pairplot.png", bbox_inches="tight")
plt.close()
print("  ✔ Saved: figures/10_pairplot.png")


# ─────────────────────────────────────────────────────────
# PLOT 11 — Metrics Heatmap — THREE SEPARATE FIGURES
# ─────────────────────────────────────────────────────────
for metric_key, cmap, fig_suffix in [("R²",   "Greens",  "a"),
                                      ("RMSE", "Reds_r",  "b"),
                                      ("MAE",  "Blues_r", "c")]:
    fig, ax = plt.subplots(figsize=(8, 4))
    heat_data = metrics_df.pivot(index="Model", columns="Property", values=metric_key)
    heat_data = heat_data.loc[["Random Forest", "XGBoost", "ANN"]]
    sns.heatmap(heat_data, annot=True, fmt=".4f", cmap=cmap, ax=ax,
                linewidths=0.5, annot_kws={"size": 10, "weight": "bold"},
                cbar_kws={"shrink": 0.8})
    ax.set_title(f"Performance Metrics Heatmap — {metric_key}",
                 fontsize=12, fontweight="bold")
    ax.set_xlabel(""); ax.set_ylabel("")
    plt.tight_layout()
    fname = f"figures/11{fig_suffix}_metrics_heatmap_{metric_key.replace('²','2')}.png"
    plt.savefig(fname, bbox_inches="tight")
    plt.close()
    print(f"  ✔ Saved: {fname}")


# ─────────────────────────────────────────────────────────
# PLOT 12 — Property trends — TWO SEPARATE FIGURES
# ─────────────────────────────────────────────────────────
# 12a: Tensile & Flexural
fig, ax = plt.subplots(figsize=(9, 5))
ax.plot(LOADINGS, mean_df["Tensile"],  "o-",  color="#2d6a4f", lw=2.2, ms=9, label="Tensile (MPa)")
ax.plot(LOADINGS, mean_df["Flexural"], "s--", color="#52b788", lw=2.2, ms=9, label="Flexural (MPa)")
ax.set_xlabel("Bamboo Powder Content (%)", fontsize=12)
ax.set_ylabel("Strength (MPa)", fontsize=11)
ax.set_xticks(LOADINGS)
ax.set_title("Effect of Bamboo Content — Tensile & Flexural Strength", fontsize=13, fontweight="bold")
ax.legend(fontsize=10)
plt.tight_layout()
plt.savefig("figures/12a_trend_Tensile_Flexural.png", bbox_inches="tight")
plt.close()
print("  ✔ Saved: figures/12a_trend_Tensile_Flexural.png")

# 12b: Hardness & Impact
fig, ax = plt.subplots(figsize=(9, 5))
ax.plot(LOADINGS, mean_df["Hardness"], "^:",  color="#e76f51", lw=2.2, ms=9, label="Hardness (Rockwell)")
ax.plot(LOADINGS, mean_df["Impact"],   "D-.", color="#457b9d", lw=2.2, ms=9, label="Impact (J/mm²)")
ax.set_xlabel("Bamboo Powder Content (%)", fontsize=12)
ax.set_ylabel("Property Value", fontsize=11)
ax.set_xticks(LOADINGS)
ax.set_title("Effect of Bamboo Content — Hardness & Impact Strength", fontsize=13, fontweight="bold")
ax.legend(fontsize=10)
plt.tight_layout()
plt.savefig("figures/12b_trend_Hardness_Impact.png", bbox_inches="tight")
plt.close()
print("  ✔ Saved: figures/12b_trend_Hardness_Impact.png")


# ─────────────────────────────────────────────────────────
# FINAL SUMMARY
# ─────────────────────────────────────────────────────────
print("\n" + "=" * 65)
print("  BEST MODEL PER PROPERTY (by R²)")
print("=" * 65)
for t in TARGETS:
    scores = {
        "Random Forest": rf_metrics[t]["R2"],
        "XGBoost":       xgb_metrics[t]["R2"],
        "ANN":           ann_metrics[t]["R2"],
    }
    best = max(scores, key=scores.get)
    print(f"  {t:<12}  →  {best:<15}  R²={scores[best]:.4f}")

print("\n" + "=" * 65)
print("  ALL OUTPUTS SAVED IN: ./figures/")
print("=" * 65)
print("""
  EXCEL WORKBOOK: bamboo_pbs_ml_results.xlsx
  ─────────────────────────────────────────────
  Sheet 1  Raw Dataset          (all 12 observations)
  Sheet 2  Mean Summary         (means + std by loading level)
  Sheet 3  Model Metrics        (R², RMSE, MAE + best model table)
  Sheet 4  LOOCV Predictions    (actual vs RF/XGB/ANN per sample)
  Sheet 5  Correlation Matrix   (Pearson, colour-coded)
  Sheet 6  Figure Index         (all saved PNG files)

  FIGURES (each graph = separate PNG file)
  ─────────────────────────────────────────────
  01a-d   Experimental bar charts (one per property)
  02a-d   ML prediction curves   (one per property)
  03a-l   Actual vs Predicted    (one per model × property)
  04      R² grouped bar comparison
  05a-b   RMSE / MAE bar comparisons
  06      Pearson correlation heatmap
  07a-l   Residuals stem plots   (one per model × property)
  08      RF feature importance
  09      Radar chart (R² per model × property)
  10      Pairplot coloured by bamboo loading
  11a-c   Metrics heatmaps       (R² / RMSE / MAE separately)
  12a-b   Property trends        (Tensile+Flexural / Hardness+Impact)
""")
